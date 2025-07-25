#!/usr/bin/env python3
"""
Invoice PDF Parser - Automated Invoice Data Extraction
=====================================================

A robust solution for extracting structured data from invoice PDFs using
multiple extraction methods including OCR, text parsing, and AI fallback.

Author: Assistant
License: MIT
"""

import os
import re
import json
import logging
import traceback
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Any
from dataclasses import dataclass, asdict
import warnings
warnings.filterwarnings("ignore")

# Core libraries
import pandas as pd
import numpy as np

# PDF processing
import fitz  # PyMuPDF
import pdfplumber
from PIL import Image
import io

# OCR
import pytesseract
from pdf2image import convert_from_path

# NLP and pattern matching
import spacy
from dateutil import parser as date_parser

# Excel output
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Optional: OpenAI for complex cases (requires API key)
try:
    import openai
    OPENAI_AVAILABLE = True
except ImportError:
    OPENAI_AVAILABLE = False

@dataclass
class InvoiceData:
    """Structure for extracted invoice data"""
    filename: str
    vendor_name: str = ""
    invoice_number: str = ""
    invoice_date: str = ""
    due_date: str = ""
    currency: str = ""
    subtotal: float = 0.0
    tax_amount: float = 0.0
    total_amount: float = 0.0
    payment_terms: str = ""
    extraction_confidence: float = 0.0
    extraction_method: str = ""
    errors: List[str] = None
    
    def __post_init__(self):
        if self.errors is None:
            self.errors = []

class InvoiceParser:
    """
    Main invoice parsing class with multiple extraction strategies
    """
    
    def __init__(self, config_path: Optional[str] = None):
        self.setup_logging()
        self.load_config(config_path)
        self.setup_nlp()
        self.setup_patterns()
        
    def setup_logging(self):
        """Configure logging"""
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler('invoice_parser.log'),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)
        
    def load_config(self, config_path: Optional[str]):
        """Load configuration settings"""
        default_config = {
            "ocr_confidence_threshold": 60,
            "use_openai_fallback": False,
            "openai_api_key": "",
            "tesseract_config": "--oem 3 --psm 6",
            "supported_currencies": ["USD", "$", "EUR", "€", "GBP", "£"],
            "date_formats": ["%m/%d/%Y", "%d/%m/%Y", "%Y-%m-%d", "%B %d, %Y"],
            "vendor_blacklist": ["COPY", "DUPLICATE", "VOID"]
        }
        
        if config_path and os.path.exists(config_path):
            try:
                with open(config_path, 'r') as f:
                    user_config = json.load(f)
                default_config.update(user_config)
            except Exception as e:
                self.logger.warning(f"Could not load config file: {e}")
                
        self.config = default_config
        
        # Setup OpenAI if available and configured
        if OPENAI_AVAILABLE and self.config["use_openai_fallback"] and self.config["openai_api_key"]:
            openai.api_key = self.config["openai_api_key"]
            self.use_openai = True
        else:
            self.use_openai = False
            
    def setup_nlp(self):
        """Initialize NLP components"""
        try:
            # Load spaCy model (download with: python -m spacy download en_core_web_sm)
            self.nlp = spacy.load("en_core_web_sm")
        except OSError:
            self.logger.warning("spaCy model not found. NLP features will be limited.")
            self.nlp = None
            
    def setup_patterns(self):
        """Define regex patterns for field extraction"""
        self.patterns = {
            'invoice_number': [
                r'(?:invoice\s*#?|inv\s*#?|invoice\s*no\.?|inv\s*no\.?)\s*:?\s*([A-Z0-9\-_]+)',
                r'#\s*([A-Z0-9\-_]{3,})',
                r'(?:^|\s)([A-Z]{2,}[0-9]{3,})',
            ],
            'vendor_name': [
                r'^([A-Z][A-Za-z\s&.,\-]+(?:Inc|LLC|Ltd|Corp|Company|Co\.?)?)(?:\n|$)',
                r'(?:from|bill\s*to|vendor):\s*([A-Za-z\s&.,\-]+)',
                r'^([A-Z\s&.,\-]+)(?:\n.*address|\n.*phone|\n.*email)',
            ],
            'amounts': [
                r'(?:total|amount\s*due|balance)\s*:?\s*\$?\s*([\d,]+\.?\d*)',
                r'(?:subtotal|sub\s*total)\s*:?\s*\$?\s*([\d,]+\.?\d*)',
                r'(?:tax|vat)\s*:?\s*\$?\s*([\d,]+\.?\d*)',
                r'\$\s*([\d,]+\.?\d*)',
            ],
            'dates': [
                r'(?:invoice\s*date|date)\s*:?\s*(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4})',
                r'(?:due\s*date|payment\s*due)\s*:?\s*(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{2,4})',
                r'(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{4})',
                r'([A-Za-z]+\s+\d{1,2},?\s+\d{4})',
            ],
            'payment_terms': [
                r'(?:terms|payment\s*terms)\s*:?\s*([^.\n]+)',
                r'(?:net\s*\d+|due\s*in\s*\d+\s*days?|payment\s*due[^.\n]+)',
            ],
            'currency': [
                r'([A-Z]{3})\s*[\d,]+\.?\d*',
                r'(USD|EUR|GBP)\b',
                r'(\$|€|£)',
            ]
        }
        
    def extract_text_from_pdf(self, pdf_path: str) -> Tuple[str, str]:
        """
        Extract text using multiple methods
        Returns: (text_content, extraction_method)
        """
        methods_tried = []
        
        # Method 1: PyMuPDF (fast, good for text-based PDFs)
        try:
            doc = fitz.open(pdf_path)
            text = ""
            for page in doc:
                text += page.get_text()
            doc.close()
            
            if len(text.strip()) > 100:
                return text, "pymupdf"
            methods_tried.append("pymupdf")
        except Exception as e:
            self.logger.debug(f"PyMuPDF failed: {e}")
            methods_tried.append(f"pymupdf_error: {str(e)[:50]}")
            
        # Method 2: pdfplumber (better for tables and structured content)
        try:
            with pdfplumber.open(pdf_path) as pdf:
                text = ""
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
                        
            if len(text.strip()) > 100:
                return text, "pdfplumber"
            methods_tried.append("pdfplumber")
        except Exception as e:
            self.logger.debug(f"pdfplumber failed: {e}")
            methods_tried.append(f"pdfplumber_error: {str(e)[:50]}")
            
        # Method 3: OCR with Tesseract
        try:
            text = self.ocr_pdf(pdf_path)
            if len(text.strip()) > 50:
                return text, "tesseract_ocr"
            methods_tried.append("tesseract_ocr")
        except Exception as e:
            self.logger.debug(f"OCR failed: {e}")
            methods_tried.append(f"ocr_error: {str(e)[:50]}")
            
        # If all methods failed
        self.logger.warning(f"All extraction methods failed for {pdf_path}: {methods_tried}")
        return "", f"failed: {', '.join(methods_tried)}"
        
    def ocr_pdf(self, pdf_path: str) -> str:
        """Perform OCR on PDF using Tesseract"""
        try:
            # Convert PDF to images
            images = convert_from_path(pdf_path, dpi=300, first_page=1, last_page=3)  # First 3 pages
            
            text = ""
            for i, image in enumerate(images):
                # Enhance image for better OCR
                image = image.convert('L')  # Convert to grayscale
                
                # OCR with Tesseract
                page_text = pytesseract.image_to_string(
                    image, 
                    config=self.config["tesseract_config"]
                )
                text += page_text + "\n"
                
            return text
        except Exception as e:
            self.logger.error(f"OCR failed for {pdf_path}: {e}")
            return ""
            
    def extract_fields(self, text: str, filename: str) -> InvoiceData:
        """Extract structured fields from text"""
        invoice_data = InvoiceData(filename=filename)
        confidence_scores = []
        
        # Clean text
        text = re.sub(r'\s+', ' ', text)  # Normalize whitespace
        text_lower = text.lower()
        
        # Extract Invoice Number
        invoice_num, conf = self.extract_invoice_number(text)
        invoice_data.invoice_number = invoice_num
        confidence_scores.append(conf)
        
        # Extract Vendor Name
        vendor, conf = self.extract_vendor_name(text)
        invoice_data.vendor_name = vendor
        confidence_scores.append(conf)
        
        # Extract Dates
        inv_date, due_date, conf = self.extract_dates(text)
        invoice_data.invoice_date = inv_date
        invoice_data.due_date = due_date
        confidence_scores.append(conf)
        
        # Extract Amounts
        subtotal, tax, total, currency, conf = self.extract_amounts(text)
        invoice_data.subtotal = subtotal
        invoice_data.tax_amount = tax
        invoice_data.total_amount = total
        invoice_data.currency = currency
        confidence_scores.append(conf)
        
        # Extract Payment Terms
        terms, conf = self.extract_payment_terms(text)
        invoice_data.payment_terms = terms
        confidence_scores.append(conf)
        
        # Calculate overall confidence
        invoice_data.extraction_confidence = np.mean([c for c in confidence_scores if c > 0])
        
        return invoice_data
        
    def extract_invoice_number(self, text: str) -> Tuple[str, float]:
        """Extract invoice number with confidence score"""
        for pattern in self.patterns['invoice_number']:
            matches = re.findall(pattern, text, re.IGNORECASE | re.MULTILINE)
            if matches:
                # Choose the most likely invoice number
                for match in matches:
                    if len(match) >= 3 and not match.lower() in ['invoice', 'inv', 'no', 'number']:
                        return match.strip(), 0.9
                        
        # Fallback: look for any alphanumeric sequence that might be an invoice number
        fallback_matches = re.findall(r'\b([A-Z0-9]{4,})\b', text)
        if fallback_matches:
            return fallback_matches[0], 0.5
            
        return "", 0.0
        
    def extract_vendor_name(self, text: str) -> Tuple[str, float]:
        """Extract vendor name with confidence score"""
        lines = text.split('\n')
        
        # Strategy 1: First non-empty line is often vendor name
        for line in lines[:5]:  # Check first 5 lines
            line = line.strip()
            if len(line) > 2 and not re.match(r'^\d+$', line):
                # Clean up the line
                cleaned = re.sub(r'[^\w\s&.,\-]', '', line)
                if len(cleaned) > 2:
                    return cleaned[:50], 0.8
                    
        # Strategy 2: Pattern matching
        for pattern in self.patterns['vendor_name']:
            matches = re.findall(pattern, text, re.IGNORECASE | re.MULTILINE)
            if matches:
                vendor = matches[0].strip()
                if len(vendor) > 2:
                    return vendor[:50], 0.7
                    
        return "", 0.0
        
    def extract_dates(self, text: str) -> Tuple[str, str, float]:
        """Extract invoice date and due date"""
        dates_found = []
        
        # Find all potential dates
        for pattern in self.patterns['dates']:
            matches = re.findall(pattern, text, re.IGNORECASE)
            for match in matches:
                try:
                    parsed_date = date_parser.parse(match, fuzzy=True)
                    dates_found.append((match, parsed_date))
                except:
                    continue
                    
        if not dates_found:
            return "", "", 0.0
            
        # Sort dates chronologically
        dates_found.sort(key=lambda x: x[1])
        
        invoice_date = ""
        due_date = ""
        
        # Assign dates based on context and chronological order
        if len(dates_found) >= 1:
            invoice_date = dates_found[0][0]
            
        if len(dates_found) >= 2:
            due_date = dates_found[-1][0]  # Latest date is likely due date
            
        confidence = 0.8 if invoice_date else 0.0
        return invoice_date, due_date, confidence
        
    def extract_amounts(self, text: str) -> Tuple[float, float, float, str, float]:
        """Extract monetary amounts and currency"""
        amounts = []
        currency = ""
        
        # Find currency
        for pattern in self.patterns['currency']:
            matches = re.findall(pattern, text, re.IGNORECASE)
            if matches:
                currency = matches[0]
                break
                
        if not currency:
            currency = "USD"  # Default assumption
            
        # Extract all monetary amounts
        amount_patterns = [
            r'\$\s*([\d,]+\.?\d*)',
            r'([\d,]+\.?\d*)\s*(?:USD|EUR|GBP|\$|€|£)',
            r'(?:total|amount|subtotal|tax)\s*:?\s*\$?\s*([\d,]+\.?\d*)',
        ]
        
        for pattern in amount_patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            for match in matches:
                try:
                    amount = float(match.replace(',', ''))
                    if amount > 0:
                        amounts.append(amount)
                except:
                    continue
                    
        if not amounts:
            return 0.0, 0.0, 0.0, currency, 0.0
            
        # Sort amounts to identify total, subtotal, tax
        amounts = sorted(set(amounts), reverse=True)
        
        total_amount = amounts[0] if amounts else 0.0
        subtotal = 0.0
        tax_amount = 0.0
        
        # Heuristic: largest amount is total, others might be subtotal/tax
        if len(amounts) >= 2:
            subtotal = amounts[1]
            
        if len(amounts) >= 3:
            tax_amount = amounts[2]
            
        # Validate relationships
        if subtotal + tax_amount > total_amount * 1.1:  # Allow 10% variance
            subtotal = total_amount * 0.85  # Estimate
            tax_amount = total_amount - subtotal
            
        confidence = 0.9 if total_amount > 0 else 0.0
        return subtotal, tax_amount, total_amount, currency, confidence
        
    def extract_payment_terms(self, text: str) -> Tuple[str, float]:
        """Extract payment terms"""
        for pattern in self.patterns['payment_terms']:
            matches = re.findall(pattern, text, re.IGNORECASE)
            if matches:
                terms = matches[0].strip()
                return terms[:100], 0.7
                
        # Look for common payment terms
        common_terms = re.findall(r'(?:net\s*\d+|due\s*in\s*\d+\s*days?)', text, re.IGNORECASE)
        if common_terms:
            return common_terms[0], 0.8
            
        return "", 0.0
        
    def process_single_pdf(self, pdf_path: str) -> InvoiceData:
        """Process a single PDF file"""
        filename = os.path.basename(pdf_path)
        self.logger.info(f"Processing: {filename}")
        
        try:
            # Extract text
            text, extraction_method = self.extract_text_from_pdf(pdf_path)
            
            if not text.strip():
                return InvoiceData(
                    filename=filename,
                    errors=["No text could be extracted from PDF"],
                    extraction_method=extraction_method
                )
                
            # Extract structured data
            invoice_data = self.extract_fields(text, filename)
            invoice_data.extraction_method = extraction_method
            
            # AI fallback for low confidence extractions
            if (invoice_data.extraction_confidence < 0.6 and 
                self.use_openai and 
                len(text) < 4000):  # OpenAI token limit consideration
                
                try:
                    ai_enhanced_data = self.enhance_with_ai(text, invoice_data)
                    if ai_enhanced_data.extraction_confidence > invoice_data.extraction_confidence:
                        invoice_data = ai_enhanced_data
                        invoice_data.extraction_method += " + openai"
                except Exception as e:
                    invoice_data.errors.append(f"AI enhancement failed: {str(e)}")
                    
            return invoice_data
            
        except Exception as e:
            self.logger.error(f"Error processing {filename}: {e}")
            return InvoiceData(
                filename=filename,
                errors=[f"Processing failed: {str(e)}"],
                extraction_method="error"
            )
            
    def enhance_with_ai(self, text: str, current_data: InvoiceData) -> InvoiceData:
        """Use OpenAI to enhance extraction for difficult cases"""
        prompt = f"""
        Extract the following information from this invoice text:
        - Vendor Name
        - Invoice Number  
        - Invoice Date
        - Due Date
        - Currency
        - Subtotal Amount
        - Tax Amount
        - Total Amount
        - Payment Terms
        
        Invoice text:
        {text[:3000]}
        
        Return as JSON with these exact keys: vendor_name, invoice_number, invoice_date, due_date, currency, subtotal, tax_amount, total_amount, payment_terms
        """
        
        response = openai.ChatCompletion.create(
            model="gpt-3.5-turbo",
            messages=[{"role": "user", "content": prompt}],
            temperature=0
        )
        
        try:
            ai_result = json.loads(response.choices[0].message.content)
            
            # Merge AI results with current data, preferring non-empty AI results
            enhanced_data = InvoiceData(filename=current_data.filename)
            enhanced_data.vendor_name = ai_result.get("vendor_name", "") or current_data.vendor_name
            enhanced_data.invoice_number = ai_result.get("invoice_number", "") or current_data.invoice_number
            enhanced_data.invoice_date = ai_result.get("invoice_date", "") or current_data.invoice_date
            enhanced_data.due_date = ai_result.get("due_date", "") or current_data.due_date
            enhanced_data.currency = ai_result.get("currency", "") or current_data.currency
            enhanced_data.subtotal = float(ai_result.get("subtotal", 0)) or current_data.subtotal
            enhanced_data.tax_amount = float(ai_result.get("tax_amount", 0)) or current_data.tax_amount
            enhanced_data.total_amount = float(ai_result.get("total_amount", 0)) or current_data.total_amount
            enhanced_data.payment_terms = ai_result.get("payment_terms", "") or current_data.payment_terms
            enhanced_data.extraction_confidence = 0.85  # Higher confidence for AI-enhanced
            
            return enhanced_data
            
        except Exception as e:
            self.logger.warning(f"AI enhancement parsing failed: {e}")
            return current_data
            
    def process_batch(self, pdf_directory: str, output_excel: str) -> Dict[str, Any]:
        """Process a batch of PDF files"""
        pdf_files = list(Path(pdf_directory).glob("*.pdf"))
        
        if not pdf_files:
            raise ValueError(f"No PDF files found in directory: {pdf_directory}")
            
        self.logger.info(f"Found {len(pdf_files)} PDF files to process")
        
        results = []
        errors = []
        
        for pdf_path in pdf_files:
            try:
                invoice_data = self.process_single_pdf(str(pdf_path))
                results.append(invoice_data)
                
                if invoice_data.errors:
                    errors.extend([f"{invoice_data.filename}: {error}" for error in invoice_data.errors])
                    
            except Exception as e:
                error_msg = f"{pdf_path.name}: {str(e)}"
                errors.append(error_msg)
                self.logger.error(error_msg)
                
        # Generate Excel output
        self.create_excel_output(results, output_excel)
        
        # Generate summary report
        summary = self.generate_summary(results, errors)
        
        return {
            "total_files": len(pdf_files),
            "successful_extractions": len([r for r in results if r.extraction_confidence > 0.5]),
            "errors": errors,
            "summary": summary,
            "results": results
        }
        
    def create_excel_output(self, results: List[InvoiceData], output_path: str):
        """Create Excel file with extracted data"""
        # Convert to DataFrame
        df_data = []
        for result in results:
            row = asdict(result)
            row['errors'] = '; '.join(row['errors']) if row['errors'] else ''
            df_data.append(row)
            
        df = pd.DataFrame(df_data)
        
        # Reorder columns as specified
        column_order = [
            'filename', 'vendor_name', 'invoice_number', 'invoice_date', 
            'due_date', 'currency', 'subtotal', 'tax_amount', 'total_amount',
            'payment_terms', 'extraction_confidence', 'extraction_method', 'errors'
        ]
        
        df = df[column_order]
        
        # Create Excel file with formatting
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoice Data"
        
        # Add headers with formatting
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
            
        # Format header row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
        wb.save(output_path)
        self.logger.info(f"Excel file created: {output_path}")
        
    def generate_summary(self, results: List[InvoiceData], errors: List[str]) -> Dict[str, Any]:
        """Generate processing summary"""
        total = len(results)
        high_confidence = len([r for r in results if r.extraction_confidence > 0.8])
        medium_confidence = len([r for r in results if 0.5 <= r.extraction_confidence <= 0.8])
        low_confidence = len([r for r in results if r.extraction_confidence < 0.5])
        
        extraction_methods = {}
        for result in results:
            method = result.extraction_method
            extraction_methods[method] = extraction_methods.get(method, 0) + 1
            
        return {
            "total_processed": total,
            "high_confidence": high_confidence,
            "medium_confidence": medium_confidence, 
            "low_confidence": low_confidence,
            "accuracy_rate": f"{(high_confidence + medium_confidence) / total * 100:.1f}%" if total > 0 else "0%",
            "extraction_methods": extraction_methods,
            "total_errors": len(errors)
        }

def main():
    """Main execution function"""
    import argparse
    
    parser = argparse.ArgumentParser(description="Invoice PDF Parser")
    parser.add_argument("input_dir", help="Directory containing PDF files")
    parser.add_argument("output_excel", help="Output Excel file path")
    parser.add_argument("--config", help="Config file path (optional)")
    
    args = parser.parse_args()
    
    # Initialize parser
    parser_instance = InvoiceParser(args.config)
    
    try:
        # Process batch
        results = parser_instance.process_batch(args.input_dir, args.output_excel)
        
        # Print summary
        print("\n" + "="*50)
        print("PROCESSING SUMMARY")
        print("="*50)
        print(f"Total files processed: {results['total_files']}")
        print(f"Successful extractions: {results['successful_extractions']}")
        print(f"Accuracy rate: {results['summary']['accuracy_rate']}")
        print(f"Errors: {len(results['errors'])}")
        
        if results['errors']:
            print("\nErrors encountered:")
            for error in results['errors'][:10]:  # Show first 10 errors
                print(f"  - {error}")
                
        print(f"\nResults saved to: {args.output_excel}")
        print("Processing complete!")
        
    except Exception as e:
        print(f"Fatal error: {e}")
        traceback.print_exc()
        return 1
        
    return 0

if __name__ == "__main__":
    exit(main())